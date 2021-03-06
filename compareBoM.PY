from os import listdir
import openpyxl

def _get_slddrw_list(directory):
    # make list of solidworks drawings
    output_list = []

    for file in listdir(directory):   # remove all files that aren't drawing files
        if file.endswith(".SLDDRW") or file.endswith(".slddrw"):
            output_list.append(file[:-7])    # adds the filename without .slddrw to list

    return output_list

def _get_pdf_list(directory):
    # make list of PDF's
    output_list = []

    for file in listdir(directory):   # remove all files that aren't pdfs
        if file.endswith(".PDF") or file.endswith(".pdf"):
            output_list.append(file[:-4])    # adds the filename without .pdf to list

    return output_list

def _get_bom(directory):
    # read BoM and make list of BoM entries
    vendor_col = 2                          # set column of vendor property
    part_num_col = 3                        # set column of part number property
    bom = openpyxl.load_workbook(directory)  # open excel file
    sheet = bom['Sheet1']                   # select the first sheet

    # find length of BoM
    num_bom_entries = 0                     # counter of number of rows to check
    while sheet.cell(row=num_bom_entries+1, column=1).value is not None:
        num_bom_entries += 1

    # Create list of parts that have a CPS vendor name
    output_list = []                        # list of parts to output
    temp_string = ''

    for i in range(1, num_bom_entries+2):
        if (sheet.cell(row=i, column=vendor_col).value is not None) and (str.lower(sheet.cell(row=i, column=vendor_col).value) == 'cps'):
            temp_string = sheet.cell(row=i, column=part_num_col).value  # creates string to remove fromatting from
            temp_string = temp_string.lstrip()                          # removes leading spaces
            temp_string = temp_string.replace("\n", "")                 # removes any carriage returns
            output_list.append(temp_string)                             # adds string to list to return

    output_list_no_duplicates = []
    [output_list_no_duplicates.append(item) for item in output_list if item not in output_list_no_duplicates]

    return output_list_no_duplicates

def _find_missing_mirrored_parts(part_list, part_substring1, part_substring2, list_name):
    print('Mirrored part check')
    for part in part_list:                # loop through list1
        if part_substring1 in part:       # see if specified substring is in list1_part
            mirrored_part = part.replace(part_substring1, part_substring2) # make mirrored part name
            if mirrored_part not in part_list:
                print(mirrored_part, 'not found in', list_name)
    print()

def _compare_lists(list1, list2, check_list):    # see if all list1 items are in list2
    print('Comparing lists')
    for file1 in list1:
        if file1 not in list2:
            print(file1, 'not found in', check_list)
    print()


def compare_bom_pdf(BoM_file, pdf_dir):
    pdf_files = _get_pdf_list(pdf_dir)
    BoMList = _get_bom(BoM_file)
    _compare_lists(pdf_files, BoMList, 'BoM')
    _compare_lists(BoMList, pdf_files, 'pdf files')
    pass

def compare_bom_slddrw(BoM_file, slddrw_dir):
    slddrw_files = _get_slddrw_list(slddrw_dir)
    BoMList = _get_bom(BoM_file)
    _compare_lists(slddrw_files, BoMList, 'BoM')
    _compare_lists(BoMList, slddrw_files, 'slddrw files')
    pass


# See if every 350 part has a corresponding 351 part and vice versa
#find_missing_mirrored_parts(BoMList, '350', '351', 'BoM')
#find_missing_mirrored_parts(BoMList, '351', '350', 'BoM')
#find_missing_mirrored_parts(pdf_files, '350', '351', 'PDF file list')
#find_missing_mirrored_parts(pdf_files, '351', '350', 'PDF file list')